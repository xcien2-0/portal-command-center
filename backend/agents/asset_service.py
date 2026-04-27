from __future__ import annotations
import os
import json
import uuid
import io
import csv
import base64
from datetime import datetime, timezone

try:
    import qrcode
    from PIL import Image, ImageDraw, ImageFont
    import barcode
    from barcode.writer import ImageWriter
    HAS_IMAGING = True
except ImportError:
    HAS_IMAGING = False

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

_DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "db", "activos.json")

# ── Categorías y regímenes ────────────────────────────────────────────────────
CATEGORIAS = {
    "mobiliario":       "Mobiliario",
    "equipo_ti":        "Equipo TI",
    "equipo_red":       "Equipo de Red",
    "vehiculo":         "Vehículo",
    "infraestructura":  "Infraestructura",
    "herramienta":      "Herramienta",
    "cpe":              "CPE Cliente",
}

REGIMENES = ["PROPIO", "RENTADO", "LEASING", "COMODATO", "COMPARTIDO"]
ESTADOS   = ["activo", "en_mantenimiento", "dado_de_baja", "extraviado"]
EMPRESAS  = ["xcien", "luminet", "wispi", "huus", "grupo"]


# ── DB helpers ────────────────────────────────────────────────────────────────
def _cargar() -> list:
    try:
        with open(_DB_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def _guardar(activos: list):
    with open(_DB_PATH, "w", encoding="utf-8") as f:
        json.dump(activos, f, indent=2, ensure_ascii=False)


# ── CRUD ──────────────────────────────────────────────────────────────────────
def registrar(
    nombre: str,
    categoria: str,
    empresa: str,
    regimen: str,
    site: str = "",
    asignado_a: str = "",
    numero_serie: str = "",
    marca: str = "",
    modelo: str = "",
    costo_mensual: float = 0.0,
    vencimiento_contrato: str = "",
    estado: str = "activo",
    notas: str = "",
) -> dict:
    activo_id = f"ACT-{str(uuid.uuid4())[:8].upper()}"
    activo = {
        "activo_id":            activo_id,
        "nombre":               nombre,
        "categoria":            categoria,
        "categoria_label":      CATEGORIAS.get(categoria, categoria),
        "empresa":              empresa.lower(),
        "regimen":              regimen.upper(),
        "site":                 site,
        "asignado_a":           asignado_a,
        "numero_serie":         numero_serie or activo_id,
        "marca":                marca,
        "modelo":               modelo,
        "costo_mensual":        costo_mensual,
        "vencimiento_contrato": vencimiento_contrato,
        "estado":               estado,
        "notas":                notas,
        "registrado_en":        datetime.now(timezone.utc).isoformat(),
        "historial":            [],
    }
    activos = _cargar()
    activos.append(activo)
    _guardar(activos)
    print(f"✅ Activo registrado: {activo_id} — {nombre} ({empresa.upper()})")
    return activo


def listar(empresa: str = None, categoria: str = None, site: str = None) -> list:
    activos = _cargar()
    if empresa:
        activos = [a for a in activos if a.get("empresa") == empresa.lower()]
    if categoria:
        activos = [a for a in activos if a.get("categoria") == categoria]
    if site:
        activos = [a for a in activos if a.get("site", "").lower() == site.lower()]
    return activos


def obtener(activo_id: str) -> dict | None:
    return next((a for a in _cargar() if a["activo_id"] == activo_id), None)


def mover(activo_id: str, nuevo_site: str, nuevo_asignado: str, motivo: str = "") -> dict:
    activos = _cargar()
    for a in activos:
        if a["activo_id"] == activo_id:
            movimiento = {
                "fecha":          datetime.now(timezone.utc).isoformat(),
                "site_anterior":  a.get("site"),
                "site_nuevo":     nuevo_site,
                "asig_anterior":  a.get("asignado_a"),
                "asig_nuevo":     nuevo_asignado,
                "motivo":         motivo,
            }
            a["historial"].append(movimiento)
            a["site"]        = nuevo_site
            a["asignado_a"]  = nuevo_asignado
            _guardar(activos)
            return a
    raise ValueError(f"Activo no encontrado: {activo_id}")


# ── Generación de etiqueta ────────────────────────────────────────────────────
def generar_etiqueta_base64(activo_id: str) -> str:
    """Genera una etiqueta PNG con barcode Code128 + QR. Devuelve base64."""
    activo = obtener(activo_id)
    if not activo:
        raise ValueError(f"Activo no encontrado: {activo_id}")

    if not HAS_IMAGING:
        raise RuntimeError("Pillow/qrcode/python-barcode no instalados")

    W, H = 600, 300
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    # Borde
    draw.rectangle([0, 0, W-1, H-1], outline="#cccccc", width=2)

    # Header
    draw.rectangle([0, 0, W, 40], fill="#0A0A0A")
    draw.text((12, 10), f"XCIEN · {activo['empresa'].upper()}", fill="white")
    draw.text((W - 160, 10), activo.get("regimen", ""), fill="#00A859")

    # Nombre del activo
    draw.text((12, 50), activo["nombre"][:45], fill="#111111")
    draw.text((12, 72), f"{activo.get('categoria_label','')}  ·  {activo.get('site','—')}  ·  {activo.get('estado','').replace('_',' ').upper()}", fill="#555555")
    if activo.get("marca") or activo.get("modelo"):
        draw.text((12, 90), f"{activo.get('marca','')} {activo.get('modelo','')}".strip(), fill="#888888")

    # Barcode Code128
    try:
        Code128 = barcode.get_barcode_class("code128")
        bc_buf = io.BytesIO()
        Code128(activo["activo_id"], writer=ImageWriter()).write(bc_buf, options={"write_text": True, "quiet_zone": 2, "module_height": 12.0, "font_size": 8, "text_distance": 3})
        bc_buf.seek(0)
        bc_img = Image.open(bc_buf).convert("RGB")
        bc_img = bc_img.resize((340, 120))
        img.paste(bc_img, (12, 115))
    except Exception as e:
        draw.text((12, 140), f"Barcode error: {e}", fill="red")

    # QR code
    try:
        qr = qrcode.QRCode(version=2, box_size=4, border=2)
        qr.add_data(json.dumps({
            "id":       activo["activo_id"],
            "nombre":   activo["nombre"],
            "empresa":  activo["empresa"],
            "regimen":  activo["regimen"],
            "site":     activo.get("site", ""),
            "serie":    activo.get("numero_serie", ""),
        }, ensure_ascii=False))
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        qr_img = qr_img.resize((140, 140))
        img.paste(qr_img, (W - 155, 110))
    except Exception as e:
        draw.text((W - 155, 140), f"QR error: {e}", fill="red")

    # Footer
    draw.text((12, 268), f"S/N: {activo.get('numero_serie', activo_id)}", fill="#aaaaaa")
    draw.text((W - 200, 268), activo.get("registrado_en", "")[:10], fill="#aaaaaa")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


def generar_etiqueta_bytes(activo_id: str) -> bytes:
    b64 = generar_etiqueta_base64(activo_id)
    return base64.b64decode(b64)


# ── Exportar CSV compatible con Odoo ─────────────────────────────────────────
# Formato: Odoo Product Import + Stock Lot (número de serie)
ODOO_FIELDS = [
    "id", "name", "default_code", "barcode", "categ_id/name",
    "type", "uom_id/name", "description", "standard_price",
    "x_empresa", "x_regimen", "x_site", "x_asignado_a",
    "x_estado", "x_notas",
]

def exportar_csv(empresa: str = None, site: str = None) -> str:
    activos = listar(empresa=empresa, site=site)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=ODOO_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for a in activos:
        writer.writerow({
            "id":              f"activo_{a['activo_id'].lower().replace('-','_')}",
            "name":            a["nombre"],
            "default_code":    a["activo_id"],
            "barcode":         a.get("numero_serie", a["activo_id"]),
            "categ_id/name":   a.get("categoria_label", a["categoria"]),
            "type":            "product",
            "uom_id/name":     "Unidades",
            "description":     f"{a.get('marca','')} {a.get('modelo','')}".strip() or a["nombre"],
            "standard_price":  a.get("costo_mensual", 0),
            "x_empresa":       a["empresa"],
            "x_regimen":       a["regimen"],
            "x_site":          a.get("site", ""),
            "x_asignado_a":    a.get("asignado_a", ""),
            "x_estado":        a.get("estado", "activo"),
            "x_notas":         a.get("notas", ""),
        })
    return buf.getvalue()


def exportar_excel(empresa: str = None, site: str = None) -> bytes:
    if not HAS_EXCEL:
        raise RuntimeError("openpyxl no instalado")
    activos = listar(empresa=empresa, site=site)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario XCIEN"

    headers = [
        "ID Activo", "Nombre", "Categoría", "Empresa", "Régimen",
        "Site", "Asignado a", "Marca", "Modelo", "N° Serie",
        "Costo Mensual", "Vencimiento Contrato", "Estado", "Registrado",
    ]
    ws.append(headers)

    # Estilo header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="0A0A0A")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    REGIMEN_COLORS = {
        "PROPIO":    "C6EFCE",
        "RENTADO":   "FFEB9C",
        "LEASING":   "BDD7EE",
        "COMODATO":  "E2EFDA",
        "COMPARTIDO":"FCE4D6",
    }

    for a in activos:
        row = [
            a["activo_id"], a["nombre"], a.get("categoria_label", ""),
            a["empresa"].upper(), a["regimen"], a.get("site", ""),
            a.get("asignado_a", ""), a.get("marca", ""), a.get("modelo", ""),
            a.get("numero_serie", ""), a.get("costo_mensual", 0),
            a.get("vencimiento_contrato", ""), a.get("estado", ""),
            a.get("registrado_en", "")[:10],
        ]
        ws.append(row)
        color = REGIMEN_COLORS.get(a["regimen"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Anchos de columna
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
