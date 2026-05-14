#!/usr/bin/env python3
"""Genera Matriz de Habilidades Field Services en formato Excel profesional."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/home/user/live-status-hub/Matriz_Habilidades_Field_Services.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────
def fill(hex6):
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

def border(style="thin", color="BBBBBB"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, size=10, color="1F1F1F", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── palette ──────────────────────────────────────────────────────────────────
HDR_DARK  = "1A3C6E"
HDR_MID   = "2E6DB4"
HDR_COMP  = "1D5E91"
GREEN_BG  = "C6EFCE"
GREEN_FG  = "276221"
RED_BG    = "FFC7CE"
RED_FG    = "9C0006"
YEL_BG    = "FFEB9C"
YEL_FG    = "9C6500"
LEG_BG    = "EEF4FF"

AREA_BG = {
    "CAE":            "EBF3FB",
    "COR":            "FFF8E7",
    "COR-FO":         "FFEFD5",
    "OPERACIONES":    "EAFBEA",
    "INFRAESTRUCTURA":"F3E5F5",
}
AREA_FG = {
    "CAE":            "1A5276",
    "COR":            "784212",
    "COR-FO":         "4A235A",
    "OPERACIONES":    "1D6A2A",
    "INFRAESTRUCTURA":"5B2333",
}

# ── data ─────────────────────────────────────────────────────────────────────
# (nombre, estado, plaza, area, rol, hab,rb,enl,wrl,red,rout,ene,elec,ops,seg,proc,cli,soft, global%)
DATA = [
    # ── NUEVO LEON ────────────────────────────────────────────────────────
    ("AUX JAIME CASTRO GARCIA",           "Nuevo León","MONTERREY","OPERACIONES",   "Supervisor",
     True, True, False,False, True, True, True, True, True, True, True, True, True,  84.62),
    ("GUILLERMO HERNANDEZ FLORES",        "Nuevo León","MONTERREY","INFRAESTRUCTURA","Field Service",
     False,False,False,False,False,False,False, True, True, True, True, True,False,  38.46),
    ("AUX JOSE ALBERTO ESPINOZA HDEZ",    "Nuevo León","MONTERREY","OPERACIONES",   "Técnico",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("JUAN JESUS PALACIOS GUZMAN",        "Nuevo León","MONTERREY","INFRAESTRUCTURA","Técnico",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("AUX ARTURO URIEL VALENCIA PEREZ",   "Nuevo León","MONTERREY","OPERACIONES",   "Técnico",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("PEDRO CONSTANTINO MAR",             "Nuevo León","MONTERREY","OPERACIONES",   "",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("JUAN RENE NUÑEZ",                   "Nuevo León","MONTERREY","OPERACIONES",   "",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("JORGE JOSABET ALVARADO CASTILLO",   "Nuevo León","MONTERREY","CAE",           "",
     True, True, True, True, True, True, True, True,False, True, True, True, True,  92.31),
    ("AUX NESTOR CHAPARRO ONTIVEROS",     "Nuevo León","MONTERREY","CAE",           "",
     True, True, True, True, True, True, True, True, True, True,False,False,False,  76.92),
    ("AUX FRANCISCO OMAR MATURINO PAJARO","Nuevo León","MONTERREY","CAE",           "",
     True, True, True, True, True, True, True,False, True, True, True, True, True,  92.31),
    ("ALEJANDRO AVILA GARCIA",            "Nuevo León","MONTERREY","COR",           "",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("MOISES COLULA LORENZO",             "Nuevo León","MONTERREY","COR",           "",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("IVAN ISRAEL SANCHEZ MONTAÑO",       "Nuevo León","MONTERREY","COR",           "",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    # ── COAHUILA ──────────────────────────────────────────────────────────
    ("FERNANDO CASTILLO HERNANDEZ",            "Coahuila","SALTILLO","CAE","Técnico",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("GUSTAVO ANTONIO VAZQUEZ ALDAZ",          "Coahuila","SALTILLO","CAE","",
     True, True, True, True, True, True, True, True,False, True, True, True, True,  92.31),
    ("ANDRES GUADALUPE GUARDADO MALDONADO",    "Coahuila","SALTILLO","CAE","",
     True, True, True, True, True, True, True, True, True,False, True, True, True,  92.31),
    ("ALEJANDRO DARIO RAMIREZ GONZALEZ",       "Coahuila","SALTILLO","CAE","",
     True, True, True, True, True, True, True,False, True, True, True, True, True,  92.31),
    ("JOSE GUADALUPE BALDERAS MENDOZA",        "Coahuila","SALTILLO","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("ALAN ALEXIS GUTIERREZ HERNANDEZ",        "Coahuila","SALTILLO","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("JUAN ANTONIO CENICEROS GUERRERO",        "Coahuila","SALTILLO","CAE","",
     False,False, True,False,False,False,False,False,False,False,False,False,False,   7.69),
    ("MARCOS ALVA PANTALEON",                  "Coahuila","SALTILLO","CAE","",
     True, True, True, True,False, True, True, True, True, True, True, True, True,  92.31),
    ("BRIAN QUINTERO CHOREÑO",                 "Coahuila","MONTERREY","COR","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("LUIS EDUARDO HIDALGO (AUXILIAR)",        "Coahuila","MONTERREY","COR","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    # ── TAMAULIPAS ────────────────────────────────────────────────────────
    ("ANGEL GUADALUPE DAVILA TORRES",     "Tamaulipas","REYNOSA",  "CAE","",
     True, True, True,False,False, True, True, True, True, True, True, True, True,  84.62),
    ("MIGUEL ANGEL FLORES HERRERA",       "Tamaulipas","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("HECTOR IVAN MARTINEZ GALICIA",      "Tamaulipas","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    # ── QUERETARO ─────────────────────────────────────────────────────────
    ("JORGE ARMANDO RODRIGUEZ CARRILLO",       "Querétaro","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("JESUS ANDRES MARROQUIN CORONA",          "Querétaro","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("ALFREDO MARTINEZ SANCHEZ (AUXILIAR)",    "Querétaro","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("CARLOS EDUARDO GUTIERREZ HERNANDEZ",     "Querétaro","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    ("CHRISTIAN ADRIAN DE AGUERO FLORES",      "Querétaro","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
    # ── JALISCO ───────────────────────────────────────────────────────────
    ("EDWIN ALEXANDER ESTRADA MIRELES (AUX)","Jalisco","MONTERREY","CAE","",
     True, True, True, True, True, True, True, True, True, True, True, True, True, 100.00),
]

COMP_HEADERS = [
    "HABILITACIÓN","RADIOBASE","ENLACE","WIRELESS","REDES",
    "ROUTING","ENERGÍA","ELECTRICIDAD","OPERACIONES","SEGURIDAD",
    "PROCESOS","CLIENTES","SOFT SKILLS",
]

TEMARIO = [
    (1,  "Seguridad y Alturas (DC3)",
         "Uso de EPP, líneas de vida y protocolos de riesgo eléctrico.",
         "Video (sesión grabada de alturas)"),
    (2,  "Sistemas de Montaje",
         "Dominio de instalaciones empotradas, tensadas y telescópicas.",""),
    (3,  "Radiofrecuencia (RF) y Alineación",
         "Alineación de enlaces (Mimosa, Ubiquiti, Cambium) y manejo de niveles dBi.",""),
    (4,  "Networking Básico",
         "Configuración de IPs, Gateways y Routers domésticos/PYME.",""),
    (5,  "Networking Avanzado",
         "VLANs, OSPF, SDWAN y enlaces troncales (Siklu / 60G / Mikrotik).",""),
    (6,  "Cableado y Demarcación",
         "Gestión de trayectorias, resguardos y conectorización profesional.",""),
    (7,  "Sistemas de Energía",
         "Instalación de paneles solares, baterías y manejo de media tensión (440 V).",""),
    (8,  "Soluciones Especiales",
         "Dominio de Starlink, cámaras de vigilancia y micro-celdas.",""),
    (9,  "Herramientas Digitales (Odoo)",
         "Registro de tareas, consumo de inventario y firmas del cliente.",""),
    (10, "Diagnóstico y Triage",
         "Capacidad de identificar fallas (Capa 1 a Capa 3) con herramientas técnicas.",""),
    (11, "Atención y Protocolo",
         "Imagen corporativa, trato al cliente y limpieza en sitio.",""),
    (12, "Mantenimiento Preventivo/Correctivo",
         "Identificación de oxidación, tensión de retenidas y limpieza de equipos.",""),
    (13, "Gestión de Inventario",
         "Control de stock en camioneta y reportes de RMA.",""),
    (14, "Validación y Cierre",
         "Pruebas de velocidad (Speedtests) y llamada de calidad con Dispatch.",""),
]

# ── build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: MATRIZ DE HABILIDADES
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "MATRIZ DE HABILIDADES"
ws.sheet_view.showGridLines = False

NCOLS = 6 + len(COMP_HEADERS) + 2  # #,nombre,estado,plaza,area,rol + 13 + nivel + global = 21

# ── title row 1 ──
last_col = get_column_letter(NCOLS)
ws.merge_cells(f"A1:{last_col}1")
c = ws["A1"]
c.value = "ENGINEER FIELD SERVICES  ·  MATRIZ DE COMPETENCIAS TÉCNICAS Y BLANDAS"
c.font  = font(bold=True, size=14, color="FFFFFF")
c.fill  = fill(HDR_DARK)
c.alignment = align("center","center")
ws.row_dimensions[1].height = 34

# ── subtitle row 2 ──
ws.merge_cells(f"A2:{last_col}2")
c = ws["A2"]
c.value = "Evaluación de 13 competencias agrupadas | Ingeniería en Campo — Mayo 2026"
c.font  = font(bold=False, size=9, color="FFFFFF", italic=True)
c.fill  = fill(HDR_MID)
c.alignment = align("center","center")
ws.row_dimensions[2].height = 18

# ── column headers row 3 ──
headers = ["#","NOMBRE","ESTADO","PLAZA","ÁREA","ROL"] + COMP_HEADERS + ["NIVEL","GLOBAL %"]

for ci, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=ci, value=h)
    if ci <= 6:
        bg, fg = HDR_MID, "FFFFFF"
    elif ci == NCOLS:
        bg, fg = HDR_DARK, "FFFFFF"
    elif ci == NCOLS - 1:
        bg, fg = "3D3D3D", "FFFFFF"
    else:
        bg, fg = HDR_COMP, "FFFFFF"
    c.fill      = fill(bg)
    c.font      = font(bold=True, size=8, color=fg)
    c.alignment = align("center","center")
    c.border    = border()

ws.row_dimensions[3].height = 52

# ── data rows ──
for ri, row in enumerate(DATA, 4):
    nombre, estado, plaza, area, rol = row[:5]
    comps   = row[5:18]       # 13 booleans
    gpct    = row[18]         # float  e.g. 84.62

    # nivel
    if gpct >= 80:
        nivel = "AVANZADO"
    elif gpct >= 60:
        nivel = "INTERMEDIO"
    else:
        nivel = "BÁSICO"

    bg_area = AREA_BG.get(area, "FFFFFF")

    # Col 1: #
    c = ws.cell(row=ri, column=1, value=ri - 3)
    c.fill = fill(bg_area); c.font = font(bold=True, size=9)
    c.alignment = align("center"); c.border = border()

    # Col 2: NOMBRE
    c = ws.cell(row=ri, column=2, value=nombre)
    c.fill = fill(bg_area); c.font = font(size=9)
    c.alignment = align("left","center"); c.border = border()

    # Col 3: ESTADO
    c = ws.cell(row=ri, column=3, value=estado)
    c.fill = fill(bg_area); c.font = font(size=9)
    c.alignment = align("center"); c.border = border()

    # Col 4: PLAZA
    c = ws.cell(row=ri, column=4, value=plaza)
    c.fill = fill(bg_area); c.font = font(size=9)
    c.alignment = align("center"); c.border = border()

    # Col 5: ÁREA
    c = ws.cell(row=ri, column=5, value=area)
    c.fill = fill(bg_area)
    c.font = Font(bold=True, size=8, color=AREA_FG.get(area,"000000"), name="Calibri")
    c.alignment = align("center"); c.border = border()

    # Col 6: ROL
    c = ws.cell(row=ri, column=6, value=rol or "—")
    c.fill = fill("F5F5F5"); c.font = font(size=8, italic=True, color="555555")
    c.alignment = align("center"); c.border = border()

    # Cols 7-19: competencies
    for ci, val in enumerate(comps, 7):
        c = ws.cell(row=ri, column=ci)
        if val:
            c.value = "✓"; c.fill = fill(GREEN_BG)
            c.font  = Font(bold=True, size=12, color=GREEN_FG, name="Calibri")
        else:
            c.value = "✗"; c.fill = fill(RED_BG)
            c.font  = Font(bold=True, size=12, color=RED_FG, name="Calibri")
        c.alignment = align("center"); c.border = border()

    # Col 20: NIVEL
    nivel_colors = {
        "AVANZADO":  (GREEN_BG, GREEN_FG),
        "INTERMEDIO":(YEL_BG,   YEL_FG),
        "BÁSICO":    (RED_BG,   RED_FG),
    }
    nbg, nfg = nivel_colors[nivel]
    c = ws.cell(row=ri, column=NCOLS - 1, value=nivel)
    c.fill = fill(nbg); c.font = font(bold=True, size=8, color=nfg)
    c.alignment = align("center"); c.border = border()

    # Col 21: GLOBAL %
    c = ws.cell(row=ri, column=NCOLS, value=gpct / 100)
    c.number_format = "0.00%"
    if gpct >= 80:
        c.fill = fill(GREEN_BG); c.font = font(bold=True, size=10, color=GREEN_FG)
    elif gpct >= 60:
        c.fill = fill(YEL_BG);   c.font = font(bold=True, size=10, color=YEL_FG)
    else:
        c.fill = fill(RED_BG);   c.font = font(bold=True, size=10, color=RED_FG)
    c.alignment = align("center"); c.border = border()

    ws.row_dimensions[ri].height = 20

# ── summary row ──
sr = len(DATA) + 4
ws.merge_cells(f"A{sr}:F{sr}")
c = ws.cell(row=sr, column=1, value="COBERTURA POR COMPETENCIA")
c.fill = fill(HDR_DARK); c.font = font(bold=True, size=9, color="FFFFFF")
c.alignment = align("center"); c.border = border()

n = len(DATA)
for ci in range(13):
    true_cnt = sum(1 for row in DATA if row[5 + ci])
    pct = true_cnt / n
    c = ws.cell(row=sr, column=7 + ci, value=pct)
    c.number_format = "0%"
    if pct >= 0.80:
        c.fill = fill(GREEN_BG); c.font = font(bold=True, size=9, color=GREEN_FG)
    elif pct >= 0.60:
        c.fill = fill(YEL_BG);   c.font = font(bold=True, size=9, color=YEL_FG)
    else:
        c.fill = fill(RED_BG);   c.font = font(bold=True, size=9, color=RED_FG)
    c.alignment = align("center"); c.border = border()

# NIVEL summary
ws.cell(row=sr, column=NCOLS - 1, value="PROMEDIO").fill = fill("3D3D3D")
ws.cell(row=sr, column=NCOLS - 1).font = font(bold=True, size=8, color="FFFFFF")
ws.cell(row=sr, column=NCOLS - 1).alignment = align("center")
ws.cell(row=sr, column=NCOLS - 1).border = border()

avg = sum(r[18] for r in DATA) / n / 100
c = ws.cell(row=sr, column=NCOLS, value=avg)
c.number_format = "0.00%"
c.fill = fill(HDR_MID); c.font = font(bold=True, size=10, color="FFFFFF")
c.alignment = align("center"); c.border = border()
ws.row_dimensions[sr].height = 22

# ── legend ──
lr = sr + 2
legend_items = [
    ("A","B", "✓  Competencia CUBIERTA",   GREEN_BG, GREEN_FG),
    ("C","D", "✗  Competencia PENDIENTE",  RED_BG,   RED_FG),
    ("E","F", "≥ 80 %  →  AVANZADO",       GREEN_BG, GREEN_FG),
    ("G","H", "60–79 %  →  INTERMEDIO",    YEL_BG,   YEL_FG),
    ("I","J", "< 60 %  →  BÁSICO",         RED_BG,   RED_FG),
]
ws.merge_cells(f"A{lr}:J{lr}")
c = ws.cell(row=lr, column=1, value="LEYENDA")
c.fill = fill("D6E4F7"); c.font = font(bold=True, size=9, color=HDR_DARK)
c.alignment = align("center"); c.border = border()
ws.row_dimensions[lr].height = 16

for ca, cb, text, bg, fg in legend_items:
    r2 = lr + 1
    ws.merge_cells(f"{ca}{r2}:{cb}{r2}")
    c = ws.cell(row=r2, column=openpyxl.utils.column_index_from_string(ca), value=text)
    c.fill = fill(bg); c.font = font(bold=True, size=9, color=fg)
    c.alignment = align("left","center"); c.border = border()
ws.row_dimensions[lr + 1].height = 16

# ── column widths ──
col_widths = [4, 38, 13, 13, 16, 14] + [12] * 13 + [13, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# freeze header + info cols
ws.freeze_panes = "G4"

# ── thick border around full table ──
from openpyxl.styles import Side as S
thick = Side(style="medium", color="1A3C6E")
for ri2 in range(1, sr + 1):
    for ci2 in range(1, NCOLS + 1):
        c = ws.cell(row=ri2, column=ci2)
        existing = c.border
        left  = thick if ci2 == 1    else existing.left
        right = thick if ci2 == NCOLS else existing.right
        top   = thick if ri2 == 1    else existing.top
        bottom= thick if ri2 == sr   else existing.bottom
        c.border = Border(left=left, right=right, top=top, bottom=bottom)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: TEMARIO DE CAPACITACIÓN
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("TEMARIO DE CAPACITACIÓN")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value = "TEMARIO DE CAPACITACIÓN — ENGINEER FIELD SERVICES"
c.font = font(bold=True, size=13, color="FFFFFF"); c.fill = fill(HDR_DARK)
c.alignment = align("center","center"); ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:D2")
c = ws2["A2"]
c.value = "14 módulos de formación técnica y habilidades blandas para habilitación del equipo de campo"
c.font = font(size=9, italic=True, color="FFFFFF"); c.fill = fill(HDR_MID)
c.alignment = align("center","center"); ws2.row_dimensions[2].height = 18

# header
for ci, h in enumerate(["#","MÓDULO / TEMA","DESCRIPCIÓN","MÉTODO / RECURSO"], 1):
    c = ws2.cell(row=3, column=ci, value=h)
    c.fill = fill(HDR_COMP); c.font = font(bold=True, size=9, color="FFFFFF")
    c.alignment = align("center","center"); c.border = border()
ws2.row_dimensions[3].height = 22

# data
alt_colors = ["F5FAFF","FFFFFF"]
for ri, (num, tema, desc, metodo) in enumerate(TEMARIO, 4):
    alt = alt_colors[(ri - 4) % 2]
    for ci, val in enumerate([num, tema, desc, metodo], 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.fill = fill(alt)
        c.font = font(bold=(ci == 1), size=9, color="1F1F1F" if ci > 1 else HDR_DARK)
        c.alignment = align("center" if ci == 1 else "left","center")
        c.border = border()
    ws2.row_dimensions[ri].height = 28

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 72
ws2.column_dimensions["D"].width = 36
ws2.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: CATEGORÍAS DE EVALUACIÓN
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("CATEGORÍAS")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:C1")
c = ws3["A1"]
c.value = "DEFINICIÓN DE LAS 13 CATEGORÍAS DE EVALUACIÓN"
c.font = font(bold=True, size=12, color="FFFFFF"); c.fill = fill(HDR_DARK)
c.alignment = align("center","center"); ws3.row_dimensions[1].height = 28

for ci, h in enumerate(["CATEGORÍA","ESTÁNDAR EVALUADO","SUBCOMPETENCIAS CLAVE"], 1):
    c = ws3.cell(row=2, column=ci, value=h)
    c.fill = fill(HDR_MID); c.font = font(bold=True, size=9, color="FFFFFF")
    c.alignment = align("center"); c.border = border()
ws3.row_dimensions[2].height = 20

cats = [
    ("HABILITACIÓN",   "Estándar de instalación PTT + Protocolo de fallas",
     "Protocolo instalación, protocolo reparación, estándar corporativo"),
    ("RADIOBASE",      "Montaje y mantenimiento de sitios (Torres, Mástiles, Aditamentos)",
     "Torres arriostradas, mástil, telescópico, adosado, puntos de anclaje, tensado"),
    ("ENLACE",         "Equipos PtP/PmP de radiofrecuencia",
     "Radios Mikrotik/Ubiquiti/Mimosa/Cambium, SIKLU, AF-60, alineación 60G/dBi"),
    ("WIRELESS",       "Redes Wi-Fi y acceso inalámbrico",
     "Access points, herramientas WiFi, mapas de calor, estudios de línea de vista"),
    ("REDES",          "Networking básico LAN/WAN",
     "Cableado estructurado, RJ-45, IPs, gateway, DHCP, Starlink"),
    ("ROUTING",        "Enrutamiento avanzado y SD-WAN",
     "Routers Edge, VLAN, OSPF, SD-WAN, firewall, DMZ, desvío de puertos"),
    ("ENERGÍA",        "Sistemas de energía renovable y backup",
     "Paneles solares, baterías, generadores, sistemas de tierras físicas"),
    ("ELECTRICIDAD",   "Electricidad baja y media tensión",
     "110 V / 220 V / 440 V, multímetro, CD y CA, acometida eléctrica, DC3"),
    ("OPERACIONES",    "Herramientas digitales y gestión de campo",
     "Odoo (tareas, inventario, firmas), levantamientos, memoria técnica, RMA"),
    ("SEGURIDAD",      "Seguridad industrial y ciberseguridad básica",
     "EPP, alturas, firewall, Hillstone, revisión de red interna cliente"),
    ("PROCESOS",       "Protocolos corporativos y documentación",
     "Protocolo instalación, protocolo fallas, estandarización, liberación servicio"),
    ("CLIENTES",       "Atención y manejo de clientes",
     "Trato al cliente, imagen corporativa, coordinación usuarios, comunicación"),
    ("SOFT SKILLS",    "Habilidades blandas y liderazgo",
     "Trabajo en equipo, gestión del tiempo, inteligencia emocional, liderazgo"),
]

for ri, (cat, estandar, subs) in enumerate(cats, 3):
    bg = GREEN_BG if (ri % 2 == 1) else "F0FFF0"
    for ci, val in enumerate([cat, estandar, subs], 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.fill = fill(bg if ci > 1 else AREA_BG.get("CAE","EBF3FB"))
        c.font = font(bold=(ci == 1), size=9, color=AREA_FG.get("CAE", HDR_DARK) if ci == 1 else "1F1F1F")
        c.alignment = align("left","center"); c.border = border()
    ws3.row_dimensions[ri].height = 26

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 48
ws3.column_dimensions["C"].width = 70

# ── save ────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"✅  Archivo generado: {OUTPUT}")
print(f"    • Hoja 1: MATRIZ DE HABILIDADES  ({len(DATA)} técnicos, {len(COMP_HEADERS)} competencias)")
print(f"    • Hoja 2: TEMARIO DE CAPACITACIÓN ({len(TEMARIO)} módulos)")
print(f"    • Hoja 3: CATEGORÍAS (definición de las 13 categorías)")
